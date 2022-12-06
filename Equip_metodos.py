
from openpyxl import Workbook

from CoolProp.CoolProp import PropsSI as Prop
# A Classe de ciclos é criada aonde todas as propriedades do fluido possam armezenar seus dados dentro de uma instância dessa classe.
# Quando é criada a instância é necessário que o contrutor passe o fluido refrigerante daquele ciclo e quantos são os pontos de interesse em que as propriedades serão medidas.


from sympy import symbols, Eq, solve
class Ciclo:
    def __init__(self,n,fluid): # n = numeros de pontos de interesse (entre os equipamentos)
        self.fluid=fluid
        self.COP = 0    
        self.COPcarnot = n 
        self.p=['Pressao (kPa):']+["-"]*n
        self.h=['Entalpia (kJ/kg):']+["-"]*n
        self.s=['Entropia (kJ/kgK):']+["-"]*n
        self.T=['Temperatura (K):']+["-"]*n
        self.x=['Titulo:']+["-"]*n
        self.y=['Fracao massica']+["-"]*n
        self.m=['Vazão mássica']+["-"]*n
        self.erro=False
        self.errorType=''
        self.wc=['Trabalho do compressor']+["-"]*n
        self.wb=['Trabalho do compressor baixa pressao']+["-"]*n
       
        
       
    # Evaporador
    def Evapout(self,i,Pl='sat',Tsa = 0, T = 0):
        # Pl pressao low | Tsa temperatura se super aquecimento
        if Pl == 'sat':
            self.p[i] = Prop('P','T',T,'Q',1,self.fluid)/1e3 # não temos o T passado no Prop, eh necessario pedir na funcao
            Pl = self.p[i]
            self.h[i] = Prop('H','P',Pl*1e3,'Q',1,self.fluid)/1e3
            self.s[i] = Prop('S','P',Pl*1e3,'Q',1,self.fluid)/1e3
        else:        
            if Tsa == 0:
                self.T[i] = Prop('T','P',Pl*1e3,'Q',1,self.fluid)
                self.h[i] = Prop('H','P',Pl*1e3,'Q',1,self.fluid)/1e3
                self.s[i] = Prop('S','P',Pl*1e3,'Q',1,self.fluid)/1e3
            else:
                self.T[i] = Prop('T','P',Pl*1e3,'Q',1,self.fluid) + Tsa
                self.h[i] = Prop('H','P',Pl*1e3,'T',self.T[i],self.fluid)/1e3
                self.s[i] = Prop('S','P',Pl*1e3,'T',self.T[i],self.fluid)/1e3
        if i == 1:
            self.p[i] = self.p[i-2] = Pl
            if self.h[i-2] != '-':
                self.q[i] = self.h[i] - self.h[i-2]
        else:
            self.p[i-1] = self.p[i] = Pl
            if self.h[i-1] != '-':
                self.q[i] = self.h[i]-self.h[i-1]
    #Condensador
    def Condout(self,i,j=0,P='sat',T='sat'):
        # P em kPa 
        # i saida 
        # j entrada
        # dados na saida do condensador
        
        k = 0
        self.x[i] = 0
        
        if P == 'sat':
            # salvando pressao igual antes e depois do condensador
            k += 1
            P = Prop('P','T',T,'Q',0,self.fluid)/1e3
            if j == 0:
                if i == 1:
                    self.p[i] = self.p[i-2] = P
                else:
                    self.p[i] = self.p[i-1] = P
            else:
                self.p[i] = self.p[j] = P
        else:
            if j == 0:
                if i == 1:
                    self.p[i] = self.p[i-2] = P
                else:
                    self.p[i] = self.p[i-1] = P
            else:
                self.p[i] = self.p[j] = P
                
        if T == 'sat':
            # definindo temperatura de saturacao em numeros
            k += 1
            T = Prop('T','P',P*1e3,'Q',0,self.fluid)
            if j ==0:
                if i == 1:
                    self.p[i-2] = self.p[i] = P
                else:
                    self.p[i-1] = self.p[i] = P
            else:
                self.p[j] = self.p[i]            
                
        if k != 0: # saida na condicao de saturacao
            self.h[i] = Prop('H','P',P*1e3,'Q',0,self.fluid)/1e3
            self.s[i] = Prop('S','P',P*1e3,'Q',0,self.fluid)/1e3
            self.T[i] = Prop('T','P',P*1e3,'Q',0,self.fluid)
        else:      # saida nao determinada previamente
            self.h[i] = Prop('H','P',P*1e3,'T',T,self.fluid)/1e3
            self.s[i] = Prop('S','P',P*1e3,'T',T,self.fluid)/1e3
            self.T[i] = T

     

    # Valvula de expansao
    def VE(self,i,P=None,j=0):
        if P != None:
            self.p[i] = P
        if j == 0:
            if i == 1:
                j = i - 2
            else:
                j = i - 1
        self.h[i] = self.h[j]
        self.y[i] = self.y[j]
        self.s[i] = Prop('S','H',self.h[i]*1e3,'P',self.p[i]*1e3,self.fluid)/1e3
        self.T[i] = Prop('T','H',self.h[i]*1e3,'P',self.p[i]*1e3,self.fluid)
        self.x[i] = Prop('Q','H',self.h[i]*1e3,'P',self.p[i]*1e3,self.fluid)

    
    

    # Condensador 
    def CondRef(self,i,Ph,Tsr='sat'):
    
        # Ph pressao high | Tsr subresfriamento
        self.p[i] = Ph
        if Tsr == 'sat':
            self.T[i] = Prop('T','P',Ph*1e3,'Q',0,self.fluid)
            self.h[i] = Prop('H','P',Ph*1e3,'Q',0,self.fluid)/1e3
            self.s[i] = Prop('S','P',Ph*1e3,'Q',0,self.fluid)/1e3
        else:
            self.T[i] = Prop('T','P',Ph*1e3,'Q',0,self.fluid) - Tsr
            self.h[i] = Prop('H','P',Ph*1e3,'T',self.T[i],self.fluid)/1e3
            self.s[i] = Prop('S','P',Ph*1e3,'T',self.T[i],self.fluid)/1e3
        
        
        if i == 1:
            self.p[i-2] = Ph
            if self.h[i-2] != '-':
                self.q[i] = self.h[i] - self.h[i-2]
        else:
            self.p[i-1] = Ph
            
            
    # Compressor
    def Compress(self,i,P,j=0,Nis=1.0):
        # Nis = eficiencia isentropica
            
            self.p[i] = P
            if self.h[i] == '-':  # usa a eficiencia isentropica
                
                if j == 0:
                    if i == 1:
                        self.s[i] = self.s[i-2]
                        self.h[i] = Prop('H','S',self.s[i]*1e3,'P',self.p[i]*1e3,self.fluid)/1e3
                        wci = self.h[i] - self.h[i-2]
                        self.wc[i] = wci/Nis
                        self.h[i] = self.h[i-2] + self.wc[i]
                        self.s[i] = Prop('S','H',self.h[i]*1e3,'P',self.p[i]*1e3,self.fluid)/1e3
                        self.T[i] = Prop('T','H',self.h[i]*1e3,'P',self.p[i]*1e3,self.fluid)
                    else:
                        self.s[i] = self.s[i-1]
                        self.h[i] = Prop('H','S',self.s[i]*1e3,'P',self.p[i]*1e3,self.fluid)/1e3
                        print (self.h[i])
                        wci = self.h[i] - self.h[i-1]
                        self.wc[i] = wci/Nis
                        self.h[i] = self.h[i-1] + self.wc[i]
                        self.s[i] = Prop('S','H',self.h[i]*1e3,'P',self.p[i]*1e3,self.fluid)/1e3
                        self.T[i] = Prop('T','H',self.h[i]*1e3,'P',self.p[i]*1e3,self.fluid)
                else:
                        self.s[i] = self.s[j]
                        self.h[i] = Prop('H','S',self.s[i]*1e3,'P',self.p[i]*1e3,self.fluid)/1e3
                        self.wc[i] = (self.h[i] - self.h[j])/Nis
                        self.h[i] = self.h[j] + self.wc[i]
                        self.s[i] = Prop('S','H',self.h[i]*1e3,'P',self.p[i]*1e3,self.fluid)/1e3
                        self.T[i] = Prop('T','H',self.h[i]*1e3,'P',self.p[i]*1e3,self.fluid)
                        
            else:  # calcula a eficiencia isentropica
                if j == 0:
                    if i == 1:
                        si = self.s[i-2]
                        hi = Prop('H','S',si*1e3,'P',self.p[i]*1e3,self.fluid)/1e3
                        wci = hi - self.h[i-2]
                        self.wc[i] = self.h[i] - self.h[i-2]
                        self.n[i] = wci/self.wc[i]
                    else:
                        si = self.s[i-1]
                        hi = Prop('H','S',si*1e3,'P',self.p[i]*1e3,self.fluid)/1e3
                        wci = hi - self.h[i-1]
                        self.wc[i] = self.h[i] - self.h[i-1]
                        self.n[i] = wci/self.wc[i]
                else:
                        si = self.s[j]
                        hi = Prop('H','S',si*1e3,'P',self.p[i]*1e3,self.fluid)/1e3
                        wci = hi - self.h[j]
                        self.wc[i] = self.h[i] - self.h[j]
                        self.n[i] = wci/self.wc[i]

    def Estado(self,i,P,T):
        self.T[i] = T
        self.p[i] = P
        self.h[i] = Prop('H','P',P*1e3,'T',T,self.fluid)/1e3
        self.s[i] = Prop('S','P',P*1e3,'T',T,self.fluid)/1e3

   

   
    def CameraMistura(self,i,l,Vflash):
        # i estado de saida, l entrada vinda do primero compressor e Vflash é o fluxo vindo da camera de flash 
        fracMass = self.x[Vflash-3]
        self.h[i] = fracMass*self.h[Vflash] + (1-fracMass)*self.h[l]
        self.p[i] = self.p[l]
        self.s[i] = Prop('S','P',self.p[i]*1e3,'H',self.h[i]*1e3,self.fluid)
        return 
    def Tflash(self,l,v,entrada,P=0):
        #entrada são uma tupla de pontos de entrada
        #l e v são os pontos de saida liquido e vapor 
        # CASO O PARAMETRO P NAO SEJA PASSADO, HAVERA UMA VERIFICAÇÃO
        # PARA IDENTIFICAR SE ALGUM DOS PONTOS POSSUI UMA PRESSÃO DEFINIDA
        # IGUALANDO-A PARA TODOS OS PONTOS DO TFLASH
        
        if P == 0:
            if self.p[l] != '-':
                ptf = self.p[l]
                for i in entrada:
                    self.p[i] = ptf
                self.p[v] = ptf
                P = ptf
            elif self.p[v] != '-':  
                ptf = self.p[v]
                for i in entrada:
                    self.p[i] = ptf
                self.p[l] = ptf
                P = ptf
            else:
                ptf = '-'
                for k in entrada:
                    if self.p[k] != '-':
                        ptf = self.p[k]
                for i in entrada:
                    self.p[i] = ptf
                self.p[v] = ptf
                self.p[l] = ptf
                P = ptf
        else:
            for k in entrada:
                
                self.p[k] = P
            self.p[l] = P
            self.p[v] = P
        # CALCULANDO OS ESTADOS DE SAIDA SATURADOS
        self.h[l]=Prop('H','P',P*1e3,'Q',0,self.fluid)/1e3
        self.x[l]=0
        self.h[v]=Prop('H','P',P*1e3,'Q',1,self.fluid)/1e3
        self.x[l]=1
        self.s[l]=Prop('S','P',P*1e3,'Q',0,self.fluid)/1e3
        self.s[v]=Prop('S','P',P*1e3,'Q',1,self.fluid)/1e3
        
        
        ## ANALISANDO QTDE DE VARIAVEIS ##
        if len(entrada) == 1:
            for i in entrada:
                self.y[l] = (self.h[i] - self.h[v])/(self.h[l]-self.h[v])
                self.y[v] = 1 - self.y[l]
                self.y[i] = 1
                if self.m[l] != '-':
                    self.m[i] = self.m[l]/self.y[l]
                    self.m[v] = self.m[i] - self.m[l]
                elif self.m[i] != '-':
                    self.m[l] = self.y[l]*self.m[i]
                    self.m[v] = self.y[v]*self.m[i]
                elif self.m[v] != '-':
                    self.m[i] = self.m[v]/self.y[v]
                    self.m[l] = self.m[i] - self.m[v]
        else:
            if self.m[l] == '-':
                if self.m[v] == '-':
                    self.m[v] = (self.h[entrada[0]]*self.m[entrada[0]] + self.h[entrada[1]]*self.m[entrada[1]] - self.h[l]*(self.m[entrada[0]]+self.m[entrada[1]]))/(self.h[v]-self.h[l])
                    self.m[l] = self.m[entrada[0]] + self.m[entrada[1]] - self.m[v]
                if self.m[entrada[0]] == '-':
                    self.m[entrada[0]] = (self.h[v]*self.m[v] - self.h[entrada[1]]*self.m[entrada[1]] + self.h[l]*(self.m[entrada[1]]-self.m[v]))/(self.h[entrada[0]]-self.h[l])
                    self.m[l] = self.m[entrada[0]] + self.m[entrada[1]] - self.m[v]
                if self.m[entrada[1]] == '-':
                    self.m[entrada[1]] = (self.h[v]*self.m[v] - self.h[entrada[0]]*self.m[entrada[0]] + self.h[l]*(self.m[entrada[0]]-self.m[v]))/(self.h[entrada[1]]-self.h[l])
                    self.m[l] = self.m[entrada[0]] + self.m[entrada[1]] - self.m[v]
            if self.m[v] == '-':
                if self.m[entrada[0]] == '-':
                    self.m[entrada[0]] = (self.h[l]*self.m[l] - self.h[entrada[1]]*self.m[entrada[1]] + self.h[v]*(self.m[entrada[1]]-self.m[l]))/(self.h[entrada[0]]-self.h[v])
                    self.m[v] = self.m[entrada[0]] + self.m[entrada[1]] - self.m[l]
                if self.m[entrada[1]] == '-':
                    self.m[entrada[1]] = (self.h[l]*self.m[l] - self.h[entrada[0]]*self.m[entrada[0]] + self.h[v]*(self.m[entrada[0]]-self.m[l]))/(self.h[entrada[1]]-self.h[v])
                    self.m[v] = self.m[entrada[0]] + self.m[entrada[1]] - self.m[l]

    def SetMass(self,i,m):
        self.m[i] = m

    def Tub(self,*kargs):
    # Recebe todos os pontos da tubulacao para definir a fracao massica e vasao massica
        for i in kargs:
            if self.y[i] != '-':
                ft = self.y[i]
                for i in kargs:
                    self.y[i] = ft 
            if self.m[i] != '-':
                mf = self.m[i]
                for i in kargs:
                    self.m[i] = mf

    

    def COP(self,l):
        trabalho = self.TrabC()
        #print ("%0.5f kW" %trabalho)
        if self.m[l] != '-':
            Cop = self.q[l]*self.m[l]/trabalho
        elif self.y[l] != '-':
            Cop = self.q[l]*self.y[l]/self.TrabC()
        else:
            Cop = self.q[l]/self.TrabC()
        
        return round(Cop,4)

    def FracaoMass(self,n):
        lista = []
        for i in range(n):
            lista.append(i+1)
            self.y[i+1]= 1/Prop('D','P',self.p[i+1]*1e3 -1000,'T',self.T[i+1],self.fluid)
   
   
    def superAqueci(self,i,Tsa,j=0):
        if Tsa == 0:
            return None
        if j == 0:
            Temp = self.T[i]
            self.T[i] = Temp + Tsa
            self.h[i] = Prop('H','T',self.T[i],'P',self.p[i]*1e3,self.fluid)/1e3
        else:
            Temp = self.T[j]
            self.T[i] = Temp + Tsa
            self.h[i] = Prop('H','T',self.T[i],'P',self.p[i]*1e3,self.fluid)/1e3
            self.s[i] = Prop('S','T',self.T[i],'P',self.p[i]*1e3,self.fluid)/1e3  
    def subResfri(self,i,Tsub,j=0):    
        if Tsub == 0:
            return None
        if j == 0:
            Temp = self.T[i]
            self.T[i] = Temp - Tsub
            self.h[i] = Prop('H','T',self.T[i],'P',self.p[i]*1e3,self.fluid)/1e3
        else:
            Temp = self.T[j]
            self.T[i] = Temp - Tsub
            self.h[i] = Prop('H','T',self.T[i],'P',self.p[i]*1e3,self.fluid)/1e3
            self.s[i] = Prop('S','T',self.T[i],'P',self.p[i]*1e3,self.fluid)/1e3  
        
    def CriaTabelaCascata(self,cicloLow):
        wb = Workbook()
        ws = wb.active     
        colunas = ['B','C','D','E','F','G']
        for c in colunas:
            ws.column_dimensions[c].width=25 
         
        ws.append(['Pontos','Pressao (kPa):','Entalpia (kJ/kg)','Entropia (kJ/kgK)','Temperatura (K)'])
        for i in range(1,len(self.h)):
            ws.append([i,round(self.p[i],2),round(self.h[i],2),round(self.s[i],4),round(self.T[i],2)])
        for i in range(1,len(cicloLow.h)):
            ws.append([i,round(cicloLow.p[i],2),round(cicloLow.h[i],2),round(cicloLow.s[i],4),round(cicloLow.T[i],2)])
        
        ws['F1'] = 'COP'
        ws['G1'] = self.COP
        ws['F2'] = 'Trabalho do compressor de alta pressão'
        ws['G2'] = self.wc
        ws['H2'] ='kW'
        ws['F3'] = 'Trabalho do compressor de baixa pressão'
        ws['G3'] = self.wb
        ws['H3'] ='kW'
        wb.save(f'Ciclo Cascata - high pressure {self.fluid} - low pressure {cicloLow.fluid} -T0-{int(self.T[1])}.xlsx')
    
    def CriaTabelas1(self,nome):
        wb = Workbook()
        ws = wb.active
        colunas = ['B','C','D','E','G','F']
        for c in colunas:
            ws.column_dimensions[c].width=25 
        ws.append(['Pontos','Pressao (kPa):','Entalpia (kJ/kg)','Entropia (kJ/kgK)','Temperatura (K)','Volume específico (m³/kg)'])
        for i in range(1,len(self.h)):
            ws.append([i,round(self.p[i],2),round(self.h[i],2),round(self.s[i],4),round(self.T[i],2),round(self.y[i],4)])
        ws['G1'] = 'COP'
        ws['H1'] = self.COP
        ws['G2'] = 'Trabalho no compressor'
        ws['H2'] = self.wc
        ws['I2'] ='kW'
        ws['G3'] = 'Vazão de refrigerante no evaporador'
        ws['H3'] = self.m[1]
        ws['I3'] = 'kg/s'
        ws['B15']= 'Fluido Refrigerante'
        ws['C15']= self.fluid
        wb.save(f'Ciclo {nome} - {self.fluid}-COP-{self.COP}.xlsx')


    def CriaTabelasFlash(self,nome):
        wb = Workbook()
        ws = wb.active
        colunas = ['B','C','D','E','F']
        for c in colunas:
            ws.column_dimensions[c].width=25 
        ws.column_dimensions['G'].width=30
        ws.append(['Pontos','Pressao (kPa):','Entalpia (kJ/kg)','Entropia (kJ/kgK)','Temperatura (K)','Volume específico (m³/kg)'])
        for i in range(1,len(self.h)):
            ws.append([i,round(self.p[i],2),round(self.h[i],2),round(self.s[i],4),round(self.T[i],2),round(self.y[i],4)])
        ws['G1'] = 'COP'
        ws['H1'] = self.COP
        ws['G3'] = 'Trabalho no compressor de alta pressao'
        ws['H3'] = self.wc
        ws['I3'] ='kW'
        ws['G4'] = 'Trabalho no compressor de baixa pressao'
        ws['H4'] = self.wb 
        ws['I4'] ='kW'      
        ws['G2'] = 'Vazão de refrigerante no evaporador'
        ws['H2'] = self.m[1]
        ws['I2'] = 'kg/s'
        ws['B15']= 'Fluido refrigerante'
        ws['C15']= self.fluid
        wb.save(f'Ciclo {nome} - {self.fluid}-COP-{self.COP}.xlsx')
   
        

    def ResultadosWc(self):
        trabalhoCompressor = float((self.h[2] - self.h[1]))
        return trabalhoCompressor 
    def ResultadosCf(self):
        Cf = float(self.h[1]-self.h[4])
        return Cf
    

    def ResultadosCop(self):
        trabalhoCompressor = self.h[2] - self.h[1]
        calorUtil = self.h[1] - self.h[4]
        COP = calorUtil/trabalhoCompressor
        return COP
    def ResultadosCarnot(self,Te,Tc):
        CopCarnot = Te/(Tc-Te)
        Neficiente = self.COP/CopCarnot
        return Neficiente


