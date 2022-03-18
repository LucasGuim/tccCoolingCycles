
from Equipamentos_MkII import *
from CoolProp.CoolProp import PropsSI as Prop
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference, 
    Series
)
import os

# Comandos da funcao main
CF_sem_tsa_e_tsr = True
WC_sem_tsa_e_tsr = False
CF_com_tsa_e_tsr = False
WC_com_tsa_e_tsr = False
criar_planilhas = True

# PARAMETROS PADRAO ------------------------------------------------

# Do equipamento em analise:
mArCondExterno = [0.28415,0.4262,0.5683,0.7103,0.85245] # vasao massica de ar externo trocando calor  no condensador
mArEvap = 0.2552 # vasao massica de ar no ambiente interno trocando calor no evaporador
capacidadeFrigorifica = 2.64
trabalhoCompressor = 0.74
efetividadeCondensador = 0.8 
efetividadeEvaporador = 0.8 # valor suposto
eficienciaIsentropica = 0.8 # valor suposto
temperaturaAmbRefrigerado = 17 + 273.15
precisao = 0.9999 # 99,99% de precisao

# Refrigerantes selecionados + R410a do equipamento selecionado
refrigerantes = ['R717','R600a', 'R290','R1234yf', 'R1234ze(E)', 'R410a','Water','R134a']

# Temperaturas externas de 30 a 40 C
temperaturasExterna =  35+273.15
                       

# Temperaturas de analise para sistemas com sub-resfriamento e superaquecimento
temperaturasExternaSaSr = [30 +273.15, 35+273.15, 40+273.15]

# ------------------------------------------------------------------

def CicloCompressaoDeVapor_1Estagio (fluido, efetividade_cond, efetividade_evap, eficiencia_is,
                                     t_amb, t_ref, v_ar_cond, v_ar_evap, capacidade_f = 0, trabalho_c = 0,
                                     t_sa = 0, t_sr = 0):
    '''
        Descricao:
            Calcula o cop do ciclo a partir da potencia frigorifica ou do trabalho do compressor
            Temperatura de sub-resfriamento e superaquecimento é opcional

        Parametros:
            fluido: Fluido refrigerante
            efetividade_cond: Efetividade do condensador
            efetividade_evap: Efetividade do evaporador
            eficiencia_is: Eficiencia isentropica do compressor
            t_amb: Temperatura do ambiente externo [K]
            t_ref: Temperatura do ambiente refrigerado [K]
            v_ar_cond: Vasao massica de ar no condensador [kg/s]
            v_ar_evap: Vasao massica de ar no evaporador [kg/s]
            capacidade_f: Capaciadade frigorifica [kW]
            trabalho_c: Trabalho compressor [kW]
            t_sa: Temperatura de superaquecimento [K]
            t_sr: Temperatura de sub resfriamento [K]

        Observacao:
            Para a efetividade do condensador e evaporador, eh considerado a temperatura de saturacao como
        a temperatura de entrada do fluido no trocador
            O programa funciona a partir de iteracoes, eh chutado um valor inicial para a temperatura de 
        condensacao, define-se todos os pontos do ciclo e ao fim a convergencia eh analisada comparando o
        quanto de calor rejeitado ao ambiente e retirado do refrigerante, para a coerencia do sistema buscamos
        a igualdade entre essas duas grandezas
    '''

    # Calculo a partir da capacidade frigorifica
    
    if capacidade_f != 0:
        ciclo = Ciclo(4, fluido)
    
        # achamos a temperatura do evaporador a partir da capacidade frigorifica, da evetividade do evaporador e
        # da temperatura do ambiente refrigerado
        Tevap = ciclo.evapRefEfetividade(1, efetividade_evap, t_ref, t_sa, v_ar_evap, capacidade_f)

        # inicio da funcao de iteracao que ira encontrar o valor de Th
        erro = 1
        Testimada = t_amb + (t_ref - Tevap) # chute inicial - valor de T3 igual a Tamb mais a
                                            # diferenca de temperaturas entrado no evaporador
        Tmax = Testimada
        Tmin = t_amb
        iteracao = 0

        while erro > (1-precisao):
            iteracao += 1

            # define a pressao alta
            pressaoH = Prop("P", "T", Testimada, "Q", 0, fluido)/1e3
            # limpa a saida do compressor da iteracao anterior
            ciclo.h[2] = '-'
            # define o ponto 2
            ciclo.Compress(2, pressaoH, j=1, Nis=eficiencia_is)                                  
            # define o ponto 3 e calcula a vasao massica
            ciclo.CondRefEfetividade(3, t_amb, Testimada-t_amb, t_sr, efetividade_cond, v_ar_cond, 2)
            # define o ponto 4
            ciclo.VE(4, ciclo.p[1], 3)

            # agora podemos calcular a capacidade frigorifica estimada
            Qestimada = ciclo.calculoEvapRefrigeracaoEfetividade(1, 4)
            erro = (Qestimada - capacidade_f)/capacidade_f

            if erro > 0: #se o valor estimado de CF for maior que o valor real devemos pegar uma temperatura menor
                Tmax = Testimada
                Testimada = (Tmax + Tmin) / 2.0
            else:
                if iteracao == 1: #caso o chute inicial seja menor do que o Treal
                    Tmin = Testimada
                    Testimada = (Tmax-t_amb)*2+t_amb
                else:
                    Tmin = Testimada
                    Testimada = (Tmax + Tmin) / 2.0

            if erro < 0:
                erro = erro * (-1)
        return ciclo, erro
    
    # Calculo a partir do trabalho do compressor
    elif trabalho_c != 0:
        ciclo = Ciclo(4, fluido)
        Th_estimada = t_amb + 5 #chute inicial para a temperatura do refrigerante no condensador
        iteracao = 1
        erro = 1

        Tmax = Th_estimada
        Tmin = t_amb

        
        while erro > (1-precisao):
            QcondExterna = v_ar_cond * Cp_ar * (Th_estimada - t_amb) * efetividade_cond
            Qevap = QcondExterna - trabalho_c
            Tl_estimada = t_ref - Qevap / (v_ar_evap * Cp_ar * efetividade_evap)

            # ponto 1
            if t_sa != 0:
                ciclo.T[1] = Tl_estimada + t_sa
                ciclo.p[1] = Prop('P', 'T', Tl_estimada, 'Q', 1, ciclo.fluid)/1e3
                ciclo.s[1] = Prop('S', 'T', ciclo.T[1], 'P', ciclo.p[1]*1e3, ciclo.fluid)/1e3
                ciclo.h[1] = Prop('H', 'T', ciclo.T[1], 'P', ciclo.p[1]*1e3, ciclo.fluid)/1e3
            else:
                ciclo.T[1] = Tl_estimada
                ciclo.p[1] = Prop('P', 'T', ciclo.T[1], 'Q', 1, ciclo.fluid)/1e3
                ciclo.s[1] = Prop('S', 'T', ciclo.T[1], 'Q', 1, ciclo.fluid)/1e3
                ciclo.h[1] = Prop('H', 'T', ciclo.T[1], 'Q', 1, ciclo.fluid)/1e3

            # ponto 3
            if t_sr != 0:
                ciclo.T[3] = Th_estimada - t_sr
                ciclo.p[3] = Prop('P', "T", Th_estimada, "Q", 0, ciclo.fluid)/1e3
                ciclo.s[3] = Prop('S', "T", ciclo.T[3], "P", ciclo.p[3]*1e3, ciclo.fluid)/1e3
                ciclo.h[3] = Prop('H', "T", ciclo.T[3], "P", ciclo.p[3]*1e3, ciclo.fluid)/1e3
            else:
                ciclo.T[3] = Th_estimada
                ciclo.p[3] = Prop('P', "T", ciclo.T[3], "Q", 0, ciclo.fluid)/1e3
                ciclo.s[3] = Prop('S', "T", ciclo.T[3], "Q", 0, ciclo.fluid)/1e3
                ciclo.h[3] = Prop('H', "T", ciclo.T[3], "Q", 0, ciclo.fluid)/1e3

            # ponto 4
            ciclo.VE(4, P=ciclo.p[1])
            
            # limpar ponto h[2] da iteracao anterior
            ciclo.h[2] = '-'

            # ponto 2
            #ciclo.saidaCompressor(2, trabalho_c, eficiencia_is)

            ciclo.Compress(2, ciclo.p[3], j=1, Nis=eficiencia_is)

            # definindo a massa
            m = trabalho_c / (ciclo.h[2] - ciclo.h[1])
            ciclo.SetMass(2, m)
            ciclo.Tub(2, 1, 3, 4)

            QcondInterna = ciclo.CalEsp(3, 2) * ciclo.m[3] * (-1)

            #QevapInterna = ciclo.calculoEvapRefrigeracaoEfetividade(1,4)

            erro = (QcondExterna - QcondInterna) / QcondInterna

            if erro > 0 :
                Tmax = Th_estimada
                Th_estimada = (Tmax + Tmin) / 2
            else:
                Tmin = Th_estimada
                if Tmax == Th_estimada:
                    Th_estimada += 5
                    Tmax = Th_estimada
                else:
                    Th_estimada = (Tmax + Tmin) / 2
            if erro < 0:
                erro = erro * -1
            
            iteracao += 1

        return ciclo, erro


def capacidadeFrigSemSaSr():
    parametros={}

    for refrigerante in refrigerantes:
        parametros[refrigerante] = []
        for mArCond in mArCondExterno:
            dados = []
            ciclo, erro = CicloCompressaoDeVapor_1Estagio(refrigerante,efetividadeCondensador, efetividadeEvaporador,
                eficienciaIsentropica, temperaturasExterna, temperaturaAmbRefrigerado, mArCond, mArEvap, 
                capacidade_f=capacidadeFrigorifica)
            ciclo.Exibir("T", 'p', 'h', 'm', 's')
            
            CF = ciclo.calculoEvapRefrigeracaoEfetividade(1,4)
            WC = ciclo.TrabC()
            COP = ciclo.COP(1)
            EP = erro*100
            dados.append(round(CF, 4))
            dados.append(round(WC, 4))
            dados.append(round(COP, 4))
            
            
            parametros[refrigerante].append(dados)

            print (refrigerante,efetividadeCondensador)
            ciclo.Exibir('T', 'p', 'h', 'm', 's')
            print("Capacidade Frigorifica %.8f" %CF)
            print("Trabalho do compressor %.10f" %WC)
            print("COP do ciclo %0.3f" %COP)
            print(" ")
    if criar_planilhas:
        criarPlanilhaSemSaSr(parametros, "Ciclo_usando_CF_sem_Tsa_e_Tsr_efetividades")


def trabalhoSemSaSr():
    parametros = {}
    for refrigerante in refrigerantes:
        parametros[refrigerante] = []
        for mArCond in mArCondExterno:
            dados = []
            ciclo, erro = CicloCompressaoDeVapor_1Estagio(refrigerante, efetividadeCondensador, 
                efetividadeEvaporador, eficienciaIsentropica, temperaturasExterna, temperaturaAmbRefrigerado, 
                mArCond, mArEvap, trabalho_c=trabalhoCompressor)
            CF = ciclo.CalEsp(1,4) * ciclo.m[1]
            WC = ciclo.TrabC()
            COP = ciclo.COP(1)
            EP = erro*100

            dados.append(round(CF, 4))
            dados.append(round(WC, 4))
            dados.append(round(COP, 4))
            parametros[refrigerante].append(dados)

            print (refrigerante, temperatura-273)
            ciclo.Exibir("T", 'p', 'h', 'm', 's')
            print ("Capacidade Frigorifica %.4f" %CF)
            print ("Trabalho do compressor %.4f" %WC)
            print ("COP do ciclo %.4f" %COP)
            print (" ")
    if criar_planilhas:
        criarPlanilhaSemSaSr(parametros, "Ciclo_usando_WC_variando_efetCondensador_sem_Tsa_e_Tsr")


def capacidadeFrigComSaSr():
    final={}
    for refrigerante in refrigerantes:
        print("Calculando para %s" %refrigerante)
        for temperatura in temperaturasExternaSaSr:
            print(temperatura-273.15, "°C")

            difSr = 1
            difSa = 1

            Tsa = 0
            Tsr = 0

            resultados={}

            while difSr > 0:
                while difSa > 0:
                    ciclo, erro = CicloCompressaoDeVapor_1Estagio(refrigerante,efetividadeCondensador,
                        efetividadeEvaporador, eficienciaIsentropica, temperatura, temperaturaAmbRefrigerado,
                        mArCond, mArEvap, capacidade_f=capacidadeFrigorifica, t_sa=Tsa, t_sr=Tsr)
                    difSa = temperaturaAmbRefrigerado - ciclo.T[1]
                    difSr = ciclo.T[3] - temperatura
                    if difSa >= 0 and difSr >= 0:
                        CF = round(ciclo.CalEsp(1,4) * ciclo.m[1], 4)
                        WC = round(ciclo.TrabC(), 4)
                        COP = round(ciclo.COP(1), 4)
                        resultados[(Tsa, Tsr)]=(CF, WC, COP)                            
                    else:
                        break
                    
                    Tsa += 1
                difSa = 1
                Tsa = 0
                Tsr += 1
            
            final[(refrigerante, temperatura)] = resultados
    
    if criar_planilhas:
        criarPlanilhaComSrSa(final, "Ciclo_usando_CF_com_Tsa_e_Tsr")
    

def TrabalhoCompComSaSr():
    final={}
    for refrigerante in refrigerantes:
        print("Calculando para %s" %refrigerante)
        for temperatura in temperaturasExternaSaSr:
            print(temperatura-273.15, "°C")

            difSr = 1
            difSa = 1

            Tsa = 0
            Tsr = 0

            resultados={}

            while difSr > 0:
                while difSa > 0:
                    ciclo, erro = CicloCompressaoDeVapor_1Estagio(refrigerante,efetividadeCondensador,
                        efetividadeEvaporador, eficienciaIsentropica, temperatura, temperaturaAmbRefrigerado,
                        mArCond, mArEvap, trabalho_c=trabalhoCompressor, t_sa=Tsa, t_sr=Tsr)
                    difSa = temperaturaAmbRefrigerado - ciclo.T[1]
                    difSr = ciclo.T[3] - temperatura
                    if difSa >= 0 and difSr >= 0:
                        CF = round(ciclo.CalEsp(1,4) * ciclo.m[1], 4)
                        WC = round(ciclo.TrabC(), 4)
                        COP = round(ciclo.COP(1), 4)
                        resultados[(Tsa, Tsr)]=(CF, WC, COP)
                        if (Tsr==0 and Tsa==1) or (Tsr==0 and Tsa == 5):
                            ciclo.Exibir('T', 'p', 'h', 'm')
                            print(COP, CF, WC)
                    else:
                        break

                    Tsa += 1
                difSa = 1
                Tsa = 0
                Tsr += 1
            
            final[(refrigerante, temperatura)] = resultados
    
    if criar_planilhas:
        criarPlanilhaComSrSa(final, "Ciclo_usando_WC_com_Tsa_e_Tsr")


def criarPlanilhaSemSaSr(DictParametros, nome):
    book = Workbook()
    sheet = book.active

    caminho = os.path.dirname(os.path.realpath(__file__))
    indice = 1
    save = caminho + os.sep + nome + str(indice) + ".xlsx"
    while os.path.exists(save):
        indice += 1
        save = caminho + os.sep + nome + str(indice) + ".xlsx"

    chaves = list(DictParametros.keys())

    # cria a planilha
    linha = -4
    for key in chaves:
        coluna = 65
        linha += 5
        contador = 0
        sheet[chr(coluna)+str(linha)] = key
        sheet[chr(coluna)+str(linha+1)] = 'Capacidade frigorifica'
        sheet[chr(coluna)+str(linha+2)] = 'Trabalho compressor'
        sheet[chr(coluna)+str(linha+3)] = 'COP do ciclo'

        for value in DictParametros[key]:
            coluna += 1
            sheet[chr(coluna)+str(linha)] = round(mArCondExterno[contador]/0.5683*100,0)
            sheet[chr(coluna)+str(linha+1)] = value[0]
            sheet[chr(coluna)+str(linha+2)] = value[1]
            sheet[chr(coluna)+str(linha+3)] = value[2]
            contador += 1

    # cria o grafico
    chartCOP = LineChart()
    chartCOP.height = 12
    chartCOP.width = 25
    chartCOP.y_axis.title = "COP"
    chartCOP.x_axis.title = "Vazão de ar externo no condensador"

    maxCol = len(mArCondExterno)+1

    start = 4
    for i in range(len(chaves)):
        valores = Reference(worksheet=sheet, min_col=2, min_row=start, max_col=maxCol)
        serie = Series(valores, title=chaves[i])
        chartCOP.append(serie)
        start += 5

    categorias = Reference(worksheet=sheet, min_col=2, min_row=1, max_col=maxCol)
    chartCOP.set_categories(categorias)

    sheet.add_chart(chartCOP, "H2")
    book.save(save)


def criarPlanilhaComSrSa(dictParams, nome):
    legenda = ["CF", "WC", "COP"]

    print("Criando planilha")
    coluna = 65 # marca o digito da letra a

    book = Workbook()
    sheet = book.active

    for item in dictParams:
        nomeSheet = item[0] + " " + str(item[1])
        sheet.title = nomeSheet
        offset = list(dictParams[item].keys())[-1][1] + 3
        
        for elem in dictParams[item]:
            for off in range(0,3,1):
                sheet[chr(coluna)+str(offset*off + 1)] = "Tsr \ Tsa " + legenda[off]

                if elem[0] == 0:
                    sheet[chr(coluna)+str(offset*off + elem[1]+2)] = elem[1]
                if elem[1] == 0:
                    sheet[chr(coluna + elem[0] + 1) + str(offset*off + 1)] = elem[0]
                sheet[chr(coluna + elem[0] + 1) + str(offset*off + elem[1] + 2)] = round(dictParams[item][elem][off], 4)
        
        sheet = book.create_sheet()

    caminho = os.path.dirname(os.path.realpath(__file__))
    indice = 1
    save = caminho + os.sep + nome + str(indice) + ".xlsx"
    while os.path.exists(save):
        indice += 1
        save = caminho + os.sep + nome + str(indice) + ".xlsx"
    
    book.save(save)
    print("Planilha criada")

    

if __name__ == "__main__":
    if CF_sem_tsa_e_tsr:
        capacidadeFrigSemSaSr()
    if WC_sem_tsa_e_tsr:
        trabalhoSemSaSr()
    if CF_com_tsa_e_tsr:
        capacidadeFrigComSaSr()
    if WC_com_tsa_e_tsr:
        TrabalhoCompComSaSr()